#!/usr/bin/env python3
"""
Generate vault/Party Armory.md from the party's unsold non-book loot.

Reads data/loot-sheet-cache.xlsx (run fetch_loot_sheet.py first) and
data/books.json (to exclude written works already in vault/library/).

Uses the local LLM to categorize items, then writes a single organized
Obsidian page with one table per category.

Usage:
  python3 scripts/generate_armory.py               # generate/update armory
  python3 scripts/generate_armory.py --dry-run     # preview item counts only
  python3 scripts/generate_armory.py --no-llm      # skip categorization, dump raw table
  python3 scripts/generate_armory.py --refresh     # re-categorize even if cache exists
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

DATA_DIR = Path("data")
CACHE_FILE = DATA_DIR / "loot-sheet-cache.xlsx"
BOOKS_JSON = DATA_DIR / "books.json"
CATEGORY_CACHE = DATA_DIR / "armory-categories.json"
DESTROYED_FILE = DATA_DIR / "destroyed-items.json"

VAULT = Path("vault")
OUTPUT_FILE = VAULT / "Party Armory.md"

SKIP_SHEETS = {"Summary", "Gatehouse", "Notes"}

DEFAULT_ENDPOINT = "http://100.76.165.94:1234"
DEFAULT_MODEL = "google/gemma-4-26b-a4b"

BATCH_SIZE = 80   # items per LLM categorization call

CATEGORIES = [
    "magic-item",
    "weapon",
    "armor",
    "consumable",
    "tool",
    "treasure",
    "misc",
]

CATEGORY_LABELS = {
    "magic-item": "Magic Items",
    "weapon":     "Weapons & Ammunition",
    "armor":      "Armor & Shields",
    "consumable": "Consumables",
    "tool":       "Tools & Equipment",
    "treasure":   "Art, Gems & Valuables",
    "misc":       "Miscellaneous",
}

CATEGORIZE_PROMPT = """/no_think
You are cataloging items held by a DFRPG Arden Vul adventuring party.
Assign each numbered item below to exactly one category:

  magic-item  — clearly magical objects: enchanted weapons/armor, wands, staves, rings,
                amulets, necklaces, orbs, crystals, relics with stated powers
  weapon      — melee weapons, ranged weapons, ammunition, quivers, weapon components
  armor       — body armor, shields, helmets, gauntlets, greaves, cloaks of protection
  consumable  — potions, single-use items, items with limited charges that are spent
  tool        — mundane tools, keys, containers, rope, lanterns, thieves' tools, rations,
                equipment without stated magic
  treasure    — coins, gems, art objects, jewelry kept as wealth/trade goods (not as
                magic items), statuettes, decorative pieces
  misc        — everything else that doesn't fit the above

Return strictly valid JSON with the item number as key and category as value. Nothing else:
{{"items": {{"1": "magic-item", "2": "weapon", "3": "tool", ...}}}}

ITEMS:
{items}
"""


# ── Utilities ──────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r'^(the|a|an)\s+', '', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def fmt_value(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        if v == int(v):
            return f"${int(v):,}"
        return f"${v:,.0f}"
    if isinstance(v, int):
        return f"${v:,}"
    s = str(v).strip()
    if s and not s.startswith("$"):
        try:
            return f"${int(float(s)):,}"
        except ValueError:
            pass
    return s or "—"


def fmt_page(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, float) and v == int(v):
        return f"p.{int(v)}"
    return f"p.{v}"


def fmt_qty(v) -> str:
    if v is None:
        return "1"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


# ── Sheet parsing (mirrors fetch_loot_sheet.py) ────────────────────────────────

def parse_sheet_date(name: str) -> Optional[str]:
    m = re.match(r"(\d{4})(\d{2})(\d{2})", name)
    if not m:
        return None
    try:
        d = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return d.isoformat()
    except ValueError:
        return None


def find_loot_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=40):
        for cell in row:
            if isinstance(cell.value, str) and "item gained" in cell.value.lower():
                return cell.row
    return None


def extract_all_items(ws, header_row: int, sheet_date: str) -> List[Dict]:
    col_map: Dict[str, int] = {}
    for cell in ws[header_row]:
        if isinstance(cell.value, str):
            col_map[cell.value.strip().lower()] = cell.column - 1

    cols = {k: col_map.get(k) for k in
            ("item gained", "qty", "page", "item value", "notes", "sold?", "owned by")}
    item_col = cols["item gained"]
    if item_col is None:
        return []

    def _get(row, col):
        if col is None or col >= len(row):
            return None
        v = row[col]
        return None if (isinstance(v, str) and v.startswith("=")) else v

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c is None for c in row):
            continue
        item_val = row[item_col] if item_col < len(row) else None
        if not isinstance(item_val, str) or not item_val.strip():
            continue
        items.append({
            "item":  item_val.strip(),
            "qty":   _get(row, cols["qty"]),
            "page":  _get(row, cols["page"]),
            "value": _get(row, cols["item value"]),
            "notes": _get(row, cols["notes"]),
            "sold":  _get(row, cols["sold?"]),
            "owner": _get(row, cols["owned by"]),
            "date":  sheet_date,
        })
    return items


def load_party_items(wb) -> List[Dict]:
    """Load all party-held, unsold items from the workbook across all dated sheets."""
    all_items: List[Dict] = []
    sheets = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        date = parse_sheet_date(name)
        if date:
            sheets.append((date, name))
    sheets.sort()

    for date, name in sheets:
        ws = wb[name]
        hr = find_loot_header_row(ws)
        if hr is None:
            continue
        for item in extract_all_items(ws, hr, date):
            owner = str(item.get("owner") or "").strip()
            sold = str(item.get("sold") or "").strip()
            if owner.lower() == "party" and sold.lower() != "yes":
                all_items.append(item)

    return all_items


# ── Book exclusion ─────────────────────────────────────────────────────────────

def load_book_normalized_titles() -> List[str]:
    """Return sorted list of normalized book titles for substring exclusion."""
    if not BOOKS_JSON.exists():
        return []
    data = json.loads(BOOKS_JSON.read_text(encoding="utf-8"))
    return sorted({b["normalized"] for b in data}, key=len, reverse=True)


def is_book(item_key: str, book_keys: List[str]) -> bool:
    """True if item and any book title are prefix-compatible (bidirectional)."""
    for bk in book_keys:
        if item_key == bk:
            return True
        # Item name has extra notes: "deeds of phagtro the westron phagtro the scout"
        if item_key.startswith(bk + " ") or item_key.startswith(bk):
            return True
        # Book key has extra notes: "deeds of phagtro the westron phagtro the scout"
        if bk.startswith(item_key + " "):
            return True
    return False


# ── Deduplication ──────────────────────────────────────────────────────────────

def deduplicate_items(items: List[Dict]) -> List[Dict]:
    """
    Group by normalized name. Keep first acquisition date, track all dates seen.
    Prefer the most descriptive (longest) name variant.
    """
    groups: Dict[str, Dict] = {}
    for item in items:
        key = normalize(item["item"])
        if key not in groups:
            groups[key] = {
                "name":   item["item"],
                "key":    key,
                "qty":    item["qty"],
                "page":   item["page"],
                "value":  item["value"],
                "notes":  item["notes"],
                "dates":  [item["date"]],
                "first":  item["date"],
            }
        else:
            g = groups[key]
            if len(item["item"]) > len(g["name"]):
                g["name"] = item["item"]
            if item["date"] < g["first"]:
                g["first"] = item["date"]
                g["qty"]   = item["qty"]
                g["page"]  = item["page"]
                g["value"] = item["value"]
                g["notes"] = item["notes"]
            if item["date"] not in g["dates"]:
                g["dates"].append(item["date"])

    result = sorted(groups.values(), key=lambda x: x["first"])
    return result


# ── LLM ───────────────────────────────────────────────────────────────────────

def call_llm(prompt: str, endpoint: str, model: str, timeout: int = 180) -> str:
    url = endpoint.rstrip("/") + "/v1/chat/completions"
    resp = requests.post(url, json={
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 2000,
        "stream": False,
        "reasoning_effort": "none",
        "enable_thinking": False,
    }, timeout=timeout)
    resp.raise_for_status()
    msg = resp.json()["choices"][0]["message"]
    return (msg.get("content") or msg.get("reasoning_content") or "").strip()


def parse_category_response(raw: str) -> Dict[str, str]:
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        obj = json.loads(cleaned)
        return {str(k): str(v).strip() for k, v in obj.get("items", {}).items()}
    except (json.JSONDecodeError, AttributeError):
        pass
    return {}


def categorize_items(items: List[Dict], endpoint: str, model: str) -> Dict[str, str]:
    """
    Categorize all items using the LLM. Returns mapping of item key → category.
    Batches into groups of BATCH_SIZE to stay within output token limits.
    """
    results: Dict[str, str] = {}
    total = len(items)

    for batch_start in range(0, total, BATCH_SIZE):
        batch = items[batch_start: batch_start + BATCH_SIZE]
        lines = []
        for i, item in enumerate(batch, start=1):
            global_num = batch_start + i
            note = f" ({item['notes']})" if item.get("notes") else ""
            lines.append(f"{global_num}. {item['name']}{note}")

        prompt = CATEGORIZE_PROMPT.format(items="\n".join(lines))
        batch_end = min(batch_start + BATCH_SIZE, total)
        print(f"  Categorizing items {batch_start + 1}–{batch_end} / {total} …",
              file=sys.stderr)

        raw = None
        try:
            raw = call_llm(prompt, endpoint, model)
        except Exception as e:
            print(f"    LLM error: {e} — retrying in two halves …", file=sys.stderr)
            # Retry with two half-batches
            mid = len(batch) // 2
            for sub_batch, sub_offset in [(batch[:mid], 0), (batch[mid:], mid)]:
                sub_lines = []
                for j, item in enumerate(sub_batch, start=1):
                    gn = batch_start + sub_offset + j
                    note = f" ({item['notes']})" if item.get("notes") else ""
                    sub_lines.append(f"{gn}. {item['name']}{note}")
                sub_prompt = CATEGORIZE_PROMPT.format(items="\n".join(sub_lines))
                try:
                    sub_raw = call_llm(sub_prompt, endpoint, model)
                    sub_map = parse_category_response(sub_raw)
                    for j, item in enumerate(sub_batch, start=1):
                        gn = str(batch_start + sub_offset + j)
                        cat = sub_map.get(gn, "misc")
                        results[item["key"]] = cat if cat in CATEGORIES else "misc"
                except Exception as e2:
                    print(f"    Sub-batch error: {e2}", file=sys.stderr)
                    for item in sub_batch:
                        results[item["key"]] = "misc"
            continue

        mapping = parse_category_response(raw)
        for i, item in enumerate(batch, start=1):
            global_num = str(batch_start + i)
            cat = mapping.get(global_num, "misc")
            if cat not in CATEGORIES:
                cat = "misc"
            results[item["key"]] = cat

    return results


# ── Markdown generation ────────────────────────────────────────────────────────

def item_row(item: Dict) -> str:
    name = item["name"].replace("|", "\\|")
    page = fmt_page(item["page"])
    value = fmt_value(item["value"])
    qty = fmt_qty(item["qty"])
    notes = str(item.get("notes") or "").replace("|", "\\|").strip() or "—"
    acquired = item["first"]
    return f"| {name} | {qty} | {page} | {value} | {acquired} | {notes} |"


def generate_armory_page(categorized: Dict[str, List[Dict]],
                          destroyed: List[Dict]) -> str:
    today = datetime.date.today().isoformat()
    total = sum(len(v) for v in categorized.values())

    lines = [
        "---",
        "title: Party Armory",
        "---",
        "",
        "# Party Armory",
        "",
        f"*Last updated: {today} · {total} items · {len(destroyed)} destroyed*",
        "",
    ]

    for cat in CATEGORIES:
        label = CATEGORY_LABELS[cat]
        items = categorized.get(cat, [])
        if not items:
            continue

        lines += [
            f"## {label}",
            "",
            "| Item | Qty | Page | Value | Acquired | Notes |",
            "|------|-----|------|-------|----------|-------|",
        ]
        for item in sorted(items, key=lambda x: x["name"].lower()):
            lines.append(item_row(item))
        lines.append("")

    if destroyed:
        lines += [
            "## Destroyed / Lost",
            "",
            "| Item | Page | Value | Destroyed | Notes |",
            "|------|------|-------|-----------|-------|",
        ]
        for d in sorted(destroyed, key=lambda x: x.get("name", "").lower()):
            name = str(d.get("name", "")).replace("|", "\\|")
            page = fmt_page(d.get("page"))
            value = fmt_value(d.get("value"))
            date = d.get("date_destroyed", "—")
            notes = str(d.get("notes", "") or "").replace("|", "\\|") or "—"
            lines.append(f"| {name} | {page} | {value} | {date} | {notes} |")
        lines.append("")

    return "\n".join(lines)


# ── Main ───────────────────────────────────────────────────────────────────────

def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Generate vault/Party Armory.md from party-held loot."
    )
    ap.add_argument("--dry-run",    action="store_true",
                    help="Print item counts by sheet, do not write output")
    ap.add_argument("--no-llm",     action="store_true",
                    help="Skip LLM categorization; dump all items in one table")
    ap.add_argument("--refresh",    action="store_true",
                    help="Re-run LLM categorization even if cache exists")
    ap.add_argument("--fix-misc",   action="store_true",
                    help="Re-run LLM on items currently labelled misc, then regenerate")
    ap.add_argument("--endpoint",   default=DEFAULT_ENDPOINT)
    ap.add_argument("--model",      default=DEFAULT_MODEL)
    args = ap.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        print("openpyxl required: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not CACHE_FILE.exists():
        print(f"No cached sheet at {CACHE_FILE}. Run fetch_loot_sheet.py first.",
              file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(CACHE_FILE, data_only=True)
    print(f"Loaded {len(wb.sheetnames)} sheets from {CACHE_FILE}", file=sys.stderr)

    book_keys = load_book_normalized_titles()
    print(f"Excluding {len(book_keys)} book titles from library", file=sys.stderr)

    raw_items = load_party_items(wb)
    print(f"Party-held unsold items: {len(raw_items)} rows", file=sys.stderr)

    # Exclude written works (use prefix matching to catch items with parenthetical notes)
    non_book_items = [i for i in raw_items if not is_book(normalize(i["item"]), book_keys)]
    print(f"After excluding books: {len(non_book_items)} rows", file=sys.stderr)

    items = deduplicate_items(non_book_items)
    print(f"After deduplication: {len(items)} unique items", file=sys.stderr)

    if args.dry_run:
        for cat_key in ["magic-item", "weapon", "armor", "consumable", "tool", "treasure", "misc"]:
            print(f"  {CATEGORY_LABELS[cat_key]}: (run without --dry-run to categorize)")
        print("\nSample items:")
        for item in items[:15]:
            print(f"  {item['name'][:70]}")
        return

    # Load destroyed items list
    destroyed: List[Dict] = []
    if DESTROYED_FILE.exists():
        destroyed = json.loads(DESTROYED_FILE.read_text(encoding="utf-8"))
        print(f"Loaded {len(destroyed)} destroyed items from {DESTROYED_FILE}",
              file=sys.stderr)

    # Exclude destroyed items from active inventory
    destroyed_keys = {normalize(d["name"]) for d in destroyed}
    items = [i for i in items if i["key"] not in destroyed_keys]
    print(f"After excluding destroyed: {len(items)} items", file=sys.stderr)

    # Categorize
    if args.no_llm:
        key_to_cat = {item["key"]: "misc" for item in items}
    elif args.fix_misc and CATEGORY_CACHE.exists():
        print(f"Re-categorizing misc items with LLM …", file=sys.stderr)
        key_to_cat = json.loads(CATEGORY_CACHE.read_text(encoding="utf-8"))
        misc_items = [i for i in items if key_to_cat.get(i["key"]) == "misc"]
        print(f"  {len(misc_items)} items currently labeled misc", file=sys.stderr)
        if misc_items:
            new_cats = categorize_items(misc_items, args.endpoint, args.model)
            key_to_cat.update(new_cats)
        # Also categorize any new items not yet in cache
        new_items = [i for i in items if i["key"] not in key_to_cat]
        if new_items:
            print(f"  {len(new_items)} new items to categorize …", file=sys.stderr)
            key_to_cat.update(categorize_items(new_items, args.endpoint, args.model))
    elif CATEGORY_CACHE.exists() and not args.refresh:
        print(f"Using cached categories from {CATEGORY_CACHE}", file=sys.stderr)
        key_to_cat = json.loads(CATEGORY_CACHE.read_text(encoding="utf-8"))
        missing = [i for i in items if i["key"] not in key_to_cat]
        if missing:
            print(f"  {len(missing)} new items not in cache, categorizing …", file=sys.stderr)
            key_to_cat.update(categorize_items(missing, args.endpoint, args.model))
    else:
        print("Categorizing with LLM …", file=sys.stderr)
        key_to_cat = categorize_items(items, args.endpoint, args.model)

    # Save category cache
    CATEGORY_CACHE.write_text(json.dumps(key_to_cat, indent=2, ensure_ascii=False),
                               encoding="utf-8")
    print(f"Category cache saved to {CATEGORY_CACHE}", file=sys.stderr)

    # Group items by category
    categorized: Dict[str, List[Dict]] = {cat: [] for cat in CATEGORIES}
    for item in items:
        cat = key_to_cat.get(item["key"], "misc")
        if cat not in CATEGORIES:
            cat = "misc"
        categorized[cat].append(item)

    # Report
    print("\nItem counts by category:", file=sys.stderr)
    for cat in CATEGORIES:
        count = len(categorized[cat])
        if count:
            print(f"  {CATEGORY_LABELS[cat]}: {count}", file=sys.stderr)

    # Write output
    page = generate_armory_page(categorized, destroyed)
    OUTPUT_FILE.write_text(page, encoding="utf-8")
    total = sum(len(v) for v in categorized.values())
    print(f"\nWrote {total} items + {len(destroyed)} destroyed to {OUTPUT_FILE}",
          file=sys.stderr)


if __name__ == "__main__":
    main()
