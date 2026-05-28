#!/usr/bin/env python3
"""
Download the Arden Vul loot tracking Google Sheet, use a local LLM to identify
and classify written works (books, letters, maps, tablets, data crystals, etc.)
in each weekly loot log, then cross-reference those titles against Discord
weekly summary files.

Outputs deduplicated JSON to data/books.json by default.

Usage:
  python3 scripts/fetch_loot_sheet.py             # download fresh + process
  python3 scripts/fetch_loot_sheet.py --cached    # use cached XLSX (data/loot-sheet-cache.xlsx)
  python3 scripts/fetch_loot_sheet.py --refresh   # force re-download even if cache exists
  python3 scripts/fetch_loot_sheet.py --unsold-only
  python3 scripts/fetch_loot_sheet.py --output /tmp/books.json
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

SHEET_ID = "1cktovt3yQRmZNP2D2sujZHIYp0Jegn4f7H_r3912SBY"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

SKIP_SHEETS = {"Summary", "Gatehouse", "Notes"}

DATA_DIR = Path("data")
CACHE_FILE = DATA_DIR / "loot-sheet-cache.xlsx"
DEFAULT_OUTPUT = DATA_DIR / "books.json"

VAULT = Path("vault")
DISCORD_NOTES_DIR = VAULT / "notes"

DEFAULT_ENDPOINT = "http://100.76.165.94:1234"
DEFAULT_MODEL = "google/gemma-4-26b-a4b"

# Titles shorter than this many significant words get exact-phrase Discord matching only
DISCORD_MIN_WORDS = 2

BOOK_PROMPT = """/no_think
You are cataloging loot from a tabletop RPG campaign set in a megadungeon called Arden Vul.
The party found the following items on {date}. Page numbers reference the adventure module.

Identify every item that is a WRITTEN WORK — something meant to be read. This includes:
  books, tomes, codices, manuscripts, spellbooks, notebooks, diaries, journals,
  letters, correspondence, scrolls (as documents, NOT spell scrolls that cast magic),
  tablets with inscribed text, maps, quires, rescripts, official documents, deeds,
  contracts, identity papers, treasure maps, copper/stone/bone plates with writing,
  data crystals (sci-fi information storage), parchments with text, and any other
  medium intended to convey written information.

Exclude: spell scrolls ("scroll of X", "mage scroll"), scroll cases, magic items
that are not primarily reading material, coins, weapons, armor, and mundane gear.

For each written work, also classify its type as one of:
  book, spellbook, letter, map, tablet, document, codex, scroll, notebook,
  diary, data-crystal, parchment, ritual-text, or other

Return strictly valid JSON and nothing else:
{{"books": [{{"name": "exact item name as listed", "type": "type"}}]}}

If nothing qualifies, return {{"books": []}}.

ITEMS ({count} total, format: [page] item name):
{items}
"""


# ── Utilities ──────────────────────────────────────────────────────────────────

def normalize_title(s: str) -> str:
    """Lowercase, strip leading articles and punctuation, collapse whitespace."""
    s = s.lower().strip()
    s = re.sub(r'^(the|a|an)\s+', '', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def significant_words(title: str) -> List[str]:
    """Words of 4+ chars that aren't common stop words."""
    stops = {'that', 'this', 'with', 'from', 'have', 'will', 'been', 'were',
             'they', 'some', 'into', 'more', 'also', 'each', 'than', 'then',
             'when', 'what', 'which', 'whom', 'your', 'their', 'there', 'about'}
    return [w for w in re.findall(r'[A-Za-z]{4,}', title) if w.lower() not in stops]


# ── Download / cache ───────────────────────────────────────────────────────────

def download_sheet(dest: Path, force: bool = False) -> None:
    if dest.exists() and not force:
        print(f"Using cached sheet: {dest}", file=sys.stderr)
        return
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("Downloading from Google Sheets …", file=sys.stderr)
    resp = requests.get(EXPORT_URL, timeout=60)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    print(f"Saved {len(resp.content):,} bytes to {dest}", file=sys.stderr)


# ── Sheet parsing ──────────────────────────────────────────────────────────────

def parse_sheet_date(name: str) -> Optional[str]:
    m = re.match(r"(\d{4})(\d{2})(\d{2})", name)
    if not m:
        return None
    try:
        d = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return d.isoformat()
    except ValueError:
        return None


def find_loot_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=40):
        for cell in row:
            if isinstance(cell.value, str) and "item gained" in cell.value.lower():
                return cell.row
    return None


def extract_all_items(ws, header_row: int) -> List[Dict]:
    col_map: Dict[str, int] = {}
    for cell in ws[header_row]:
        if isinstance(cell.value, str):
            col_map[cell.value.strip().lower()] = cell.column - 1

    cols = {
        k: col_map.get(k)
        for k in ("item gained", "qty", "page", "item value", "notes", "sold?", "owned by")
    }
    item_col = cols["item gained"]
    if item_col is None:
        return []

    def _get(row, col):
        if col is None or col >= len(row):
            return None
        v = row[col]
        return None if (isinstance(v, str) and v.startswith("=")) else v

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c is None for c in row):
            continue
        item = row[item_col] if item_col < len(row) else None
        if not isinstance(item, str) or not item.strip():
            continue
        items.append({
            "item": item.strip(),
            "qty":   _get(row, cols["qty"]),
            "page":  _get(row, cols["page"]),
            "value": _get(row, cols["item value"]),
            "notes": _get(row, cols["notes"]),
            "sold":  _get(row, cols["sold?"]),
            "owner": _get(row, cols["owned by"]),
        })
    return items


# ── LLM ───────────────────────────────────────────────────────────────────────

def call_llm(prompt: str, endpoint: str, model: str, timeout: int = 120) -> str:
    url = endpoint.rstrip("/") + "/v1/chat/completions"
    resp = requests.post(url, json={
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 1500,
        "stream": False,
        "reasoning_effort": "none",
        "enable_thinking": False,
    }, timeout=timeout)
    resp.raise_for_status()
    msg = resp.json()["choices"][0]["message"]
    return (msg.get("content") or msg.get("reasoning_content") or "").strip()


def parse_llm_books(raw: str) -> List[Dict]:
    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        obj = json.loads(cleaned)
        items = obj.get("books", [])
        if isinstance(items, list):
            out = []
            for item in items:
                if isinstance(item, dict) and item.get("name"):
                    out.append({"name": str(item["name"]).strip(),
                                "type": str(item.get("type", "other")).strip()})
                elif isinstance(item, str) and item.strip():
                    out.append({"name": item.strip(), "type": "other"})
            return out
    except (json.JSONDecodeError, AttributeError):
        pass
    return []


def identify_books_with_llm(sheet_date: str, items: List[Dict],
                             endpoint: str, model: str) -> List[Dict]:
    item_lines = "\n".join(
        f"  [{r['page']}] {r['item']}" if r.get("page") else f"  {r['item']}"
        for r in items
    )
    prompt = BOOK_PROMPT.format(date=sheet_date, count=len(items), items=item_lines)
    try:
        raw = call_llm(prompt, endpoint, model)
    except Exception as e:
        print(f"    LLM error: {e}", file=sys.stderr)
        return []
    return parse_llm_books(raw)


# ── Discord cross-reference ────────────────────────────────────────────────────

def load_discord_summaries() -> List[Dict]:
    summaries = []
    for f in sorted(DISCORD_NOTES_DIR.glob("Discord Summary *.md")):
        text = f.read_text(encoding="utf-8", errors="ignore")
        m = re.search(
            r"\*{0,2}Date Range[:：]\*{0,2}\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})",
            text,
        )
        summaries.append({
            "week": f.stem,
            "date_start": m.group(1) if m else None,
            "date_end":   m.group(2) if m else None,
            "text": text,
        })
    return summaries


def find_in_discord(title: str, summaries: List[Dict]) -> List[str]:
    """
    Phrase-first matching:
    - Try exact phrase (case-insensitive) for titles with 2+ significant words.
    - Fall back to all-significant-words proximity check (within 200 chars).
    - Skip generic single-word titles entirely.
    """
    sig = significant_words(title)
    if not sig:
        return []

    hits = []

    if len(sig) >= DISCORD_MIN_WORDS:
        # Try exact phrase first (most reliable for named works)
        phrase_pat = re.compile(re.escape(title), re.I)
        for s in summaries:
            if phrase_pat.search(s["text"]):
                hits.append(s["week"])
        if hits:
            return hits

        # Proximity fallback: all sig words within 200 chars of the first word
        first_pat = re.compile(re.escape(sig[0]), re.I)
        rest_pats = [re.compile(re.escape(w), re.I) for w in sig[1:]]
        for s in summaries:
            text = s["text"]
            for m in first_pat.finditer(text):
                window = text[max(0, m.start() - 100): m.end() + 200]
                if all(p.search(window) for p in rest_pats):
                    hits.append(s["week"])
                    break
    # Single significant word: skip to avoid noise

    return hits


# ── Deduplication ──────────────────────────────────────────────────────────────

def deduplicate(records: List[Dict]) -> List[Dict]:
    """
    Merge acquisition records that refer to the same written work.
    Groups by normalized title, keeps the most common type, merges acquisitions
    and discord mentions.
    """
    groups: Dict[str, Dict] = {}
    for rec in records:
        key = normalize_title(rec["name"])
        if key not in groups:
            groups[key] = {
                "title":        rec["name"],
                "normalized":   key,
                "type":         rec["type"],
                "acquisitions": [],
                "discord_mentions": [],
            }
        g = groups[key]
        # Prefer the longer/more-descriptive name
        if len(rec["name"]) > len(g["title"]):
            g["title"] = rec["name"]
        # Acquisition record (one per sheet entry)
        acq = {
            "date":  rec["date"],
            "sheet": rec["sheet"],
            "qty":   rec.get("qty"),
            "page":  rec.get("page"),
            "value": rec.get("value"),
            "notes": rec.get("notes"),
            "sold":  rec.get("sold"),
            "owner": rec.get("owner"),
        }
        g["acquisitions"].append(acq)
        # Merge discord mentions (unique)
        existing = set(g["discord_mentions"])
        for w in rec.get("discord_mentions", []):
            if w not in existing:
                g["discord_mentions"].append(w)
                existing.add(w)

    result = sorted(groups.values(), key=lambda x: x["acquisitions"][0]["date"])
    return result


# ── Main ───────────────────────────────────────────────────────────────────────

def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Identify written works in the Arden Vul loot sheet using an LLM."
    )
    ap.add_argument("--cached",      action="store_true",
                    help="Use cached XLSX without downloading (data/loot-sheet-cache.xlsx)")
    ap.add_argument("--refresh",     action="store_true",
                    help="Force re-download even if cache exists")
    ap.add_argument("--unsold-only", action="store_true",
                    help="Only include items not marked as sold")
    ap.add_argument("--no-discord",  action="store_true",
                    help="Skip Discord cross-reference")
    ap.add_argument("--output",      default=str(DEFAULT_OUTPUT),
                    help=f"Output JSON path (default: {DEFAULT_OUTPUT})")
    ap.add_argument("--endpoint",    default=DEFAULT_ENDPOINT)
    ap.add_argument("--model",       default=DEFAULT_MODEL)
    args = ap.parse_args(argv)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    download_sheet(CACHE_FILE, force=args.refresh)

    try:
        import openpyxl
    except ImportError:
        print("openpyxl required: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(CACHE_FILE, data_only=True)
    print(f"Loaded {len(wb.sheetnames)} sheets", file=sys.stderr)

    discord_summaries: List[Dict] = []
    if not args.no_discord:
        discord_summaries = load_discord_summaries()
        print(f"Loaded {len(discord_summaries)} Discord summaries", file=sys.stderr)

    # Collect sheets chronologically
    sheets: List[Tuple[str, str, List[Dict]]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        sheet_date = parse_sheet_date(sheet_name)
        if not sheet_date:
            print(f"  Skipping (no date): {sheet_name}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        header_row = find_loot_header_row(ws)
        if header_row is None:
            print(f"  Skipping (no loot header): {sheet_name}", file=sys.stderr)
            continue
        items = extract_all_items(ws, header_row)
        if items:
            sheets.append((sheet_name, sheet_date, items))
    sheets.sort(key=lambda x: x[1])

    # Process each sheet
    flat_records: List[Dict] = []
    for sheet_name, sheet_date, items in sheets:
        if args.unsold_only:
            items = [i for i in items if i.get("sold") != "Yes"]
        print(f"\n{sheet_name} ({sheet_date}): {len(items)} items → LLM …", file=sys.stderr)
        books = identify_books_with_llm(sheet_date, items, args.endpoint, args.model)
        if not books:
            print("  → none", file=sys.stderr)
            continue

        print(f"  → {len(books)}: {', '.join(b['name'][:40] for b in books[:3])}"
              f"{'…' if len(books) > 3 else ''}", file=sys.stderr)

        # Build lookup by item name (strip to match LLM output)
        item_meta = {r["item"]: r for r in items}

        for b in books:
            meta = item_meta.get(b["name"], {})
            sold = meta.get("sold")
            if args.unsold_only and sold == "Yes":
                continue

            discord_hits: List[str] = []
            if discord_summaries:
                discord_hits = find_in_discord(b["name"], discord_summaries)

            flat_records.append({
                "name":             b["name"],
                "type":             b["type"],
                "date":             sheet_date,
                "sheet":            sheet_name,
                "qty":              meta.get("qty"),
                "page":             meta.get("page"),
                "value":            meta.get("value"),
                "notes":            meta.get("notes"),
                "sold":             sold,
                "owner":            meta.get("owner"),
                "discord_mentions": discord_hits,
            })
            if discord_hits:
                print(f"    '{b['name'][:50]}' → Discord: {discord_hits}", file=sys.stderr)

    # Deduplicate
    books_out = deduplicate(flat_records)

    print(f"\n{'─'*60}", file=sys.stderr)
    print(f"Unique written works: {len(books_out)}", file=sys.stderr)
    print(f"With Discord mentions: {sum(1 for b in books_out if b['discord_mentions'])}", file=sys.stderr)
    print(f"Unsold / party-held: {sum(1 for b in books_out if any(a.get('owner') == 'Party' for a in b['acquisitions']))}", file=sys.stderr)

    out_json = json.dumps(books_out, indent=2, ensure_ascii=False)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(out_json, encoding="utf-8")
    print(f"Written to {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
